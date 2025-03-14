from rest_framework.views import APIView
from rest_framework.generics import RetrieveAPIView, ListAPIView, RetrieveUpdateDestroyAPIView,ListCreateAPIView ,RetrieveUpdateAPIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.decorators import api_view
from rest_framework import status, viewsets
from rest_framework.filters import OrderingFilter
from rest_framework.decorators import action
from django.contrib.auth.models import User
from django.conf import settings
from django.template.loader import render_to_string
from django.forms.models import model_to_dict
from django.contrib.auth.hashers import make_password, check_password
from django.db.models import Q, Value, OuterRef, Subquery, F, Window, Value, CharField
from django.db.models.functions import RowNumber
from django.db import connection, transaction
from .models import Area, Departments, ModulePermissions, Modules, Roles, Employee, AccessKey, BomMasterlist, PosItems, Sales, UploadedFile, InventoryCode, BosItems, EndingInventory, Forecast, ByRequest, ByRequestItems, DeliveryCode, DeliveryItems, SalesReport, InitialReplenishment, Status, UserDefinedVariables

from .serializers import ModuleSerializer, EmployeeSerializer, AreaSerializer, RoleSerializer, DepartmentsSerializer, ChangePasswordSerializer, EmployeeSerializerPlain, RolesSerializerPlain, AccessKeySerializer, UserDetailSerializer,  ModulesSerializerParent, FileUploadSerializer, InventoryCodeSerializer, ForecastSerializer, DeliveryItemsSerializer,ByRequestSerializer, EndingInventorySerializer, SalesReportSerializer, InitialReplenishmentSerializer, ByRequestSerializerC, ByRequestItemsSerializer, UserDefinedVariablesSerializer

from collections import defaultdict
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.auth import authenticate
from django.utils import timezone, html

from django.core.mail import send_mail, EmailMultiAlternatives
from django.http import StreamingHttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
import openpyxl

from openpyxl.styles import Font, PatternFill
import logging
import random
import string
from django.core.paginator import Paginator
import pandas as pd
import math
import os
import glob
import numpy as np
import hashlib
from rapidfuzz import process


logger = logging.getLogger(__name__)


def send_approval_email(mrp, file_name):

    try:
        inventory = InventoryCode.objects.get(id=mrp.id)

        user_email = inventory.uploaded_by.email if inventory.uploaded_by else None

        if not user_email:
            logger.warning(f"No uploaded_by user for Inventory ID {mrp}")
            return
        user_email = 'npvico@gmail.com'  # Ensure `requested_by` is a User ForeignKey
        excel_link = f"http://localhost:5173/media/sales/{file_name}"

        email_subject = "MRP Approved Notification"
        email_html_content = render_to_string("emails/mrp_approved.html", {
            "user_name": mrp.uploaded_by.get_full_name(),
            "inventory_code": mrp.inventory_code,
            "excel_link": excel_link,
        })
        email_text_content = html.strip_tags(email_html_content)  # Fallback plain text email

        #create email
        email = EmailMultiAlternatives(
            subject=email_subject,
            body=email_text_content,
            from_email="rfcaguioa@gmail.com",
            to=[user_email]
        )
        email.attach_alternative(email_html_content, "text/html")
        email.send()

        logger.info(f"Approval email sent to {user_email} for MRP {mrp.id}")

    except Exception as e:
        logger.exception(f"Error sending approval email for MRP {mrp.id}: {e}")


@api_view(["POST"])
def approve_mrp(request, idofinventory):
    try:
        mrp = InventoryCode.objects.get(id=idofinventory)

        delivery_code = DeliveryCode.objects.filter(inventory_code=mrp).select_related("requested_by").first()
        mrp.status = Status.objects.get(id=5)
        mrp.approved_at=timezone.now()
        mrp.approved_by=request.user
        mrp.save()
        file_name=update_sales_report(mrp,delivery_code)
        send_approval_email(mrp, file_name)

        return Response({"success": True, "message": "MRP approved successfully."})
    except InventoryCode.DoesNotExist:
        logger.error(f"MRP with ID {idofinventory} not found.")
        return Response({"success": False, "message": "MRP not found."}, status=404)
    except Exception as e:
        logger.exception(f"Unexpected error approving MRP {idofinventory}: {e}")
        return Response({"success": False, "message": "An error occurred."}, status=500)



@api_view(["POST"])
def ammend_mrp(request, idofinventory):
    try:
        ammend_status = Status.objects.get(id=3)
        inventory = InventoryCode.objects.get(id=idofinventory)
        inventory.status = ammend_status
        inventory.save()

        delivery_code = DeliveryCode.objects.filter(inventory_code=inventory).first()
        user_message = request.data.get("message", "No additional details provided.")

        if delivery_code and delivery_code.requested_by and delivery_code.requested_by.email:
            send_mail(
                subject="MRP Amendment Notification",
                message=f"""Hello {delivery_code.requested_by.username},\n
                    Your MRP request with Inventory ID {inventory.id} needs amendment.\n\n
                    Message: {user_message}\n
                    Please take the necessary action.\n\n
                    Best Regards,\nYour Team""",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[delivery_code.requested_by.email],
                fail_silently=False,
            )

        return Response({"message": "Inventory status updated and email sent"}, status=status.HTTP_200_OK)

    except InventoryCode.DoesNotExist:
        return Response({"error": "Inventory not found"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



def update_sales_report(mrp, delivery_code):

    try:
        inventory_code_name = mrp.inventory_code  # Adjust field name if needed
        file_name = f"{inventory_code_name}_sales_report.xlsx"

        file_path = os.path.join(settings.SALES_FILES_ROOT, file_name)

        if not os.path.exists(file_path):
            logger.warning(f"⚠️ File {file_name} not found in {settings.SALES_FILES_ROOT}. Skipping update.")
            return
        delivery_items_data = list(
            DeliveryItems.objects.filter(delivery_code=delivery_code)
            .values("bom_entry__bos_code", "bom_entry__bos_material_description","bom_entry__category","bom_entry__delivery_uom",
                    "first_adjustment", "second_adjustment", "third_adjustment",
                    "first_final_delivery", "second_final_delivery", "third_final_delivery",
                    "first_qty_delivery", "second_qty_delivery", "third_qty_delivery", "first_qty_uom", "second_qty_uom", "third_qty_uom")
        )


        by_request_items_data = list(
            ByRequest.objects.filter(delivery_code=delivery_code)
            .values("by_request_item__bos_code", "by_request_item__bos_material_description","by_request_item__category","by_request_item__delivery_uom",
                    "total_weekly_request",
                    "first_delivery", "second_delivery", "third_delivery",
                    "first_final_delivery", "second_final_delivery", "third_final_delivery",
                    "first_qty_delivery", "second_qty_delivery", "third_qty_delivery", "first_qty_uom", "second_qty_byrequest_uom", "third_qty_byrequest_uom")
        )


        delivery_items_df = pd.DataFrame(delivery_items_data)
        by_request_items_df = pd.DataFrame(by_request_items_data)

        delivery_items_df.rename(columns={
            "bom_entry__bos_code": "BOS Code",
            'bom_entry__category': "Category",
            "bom_entry__bos_material_description": "Material Description",
            "bom_entry__delivery_uom" : "Delivery UOM",
            "first_adjustment": "Adjustment 1",
            "second_adjustment": "Adjustment 2",
            "third_adjustment": "Adjustment 3",
            "first_final_delivery": "Final Delivery 1",
            "second_final_delivery": "Final Delivery 2",
            "third_final_delivery": "Final Delivery 3",
            "first_qty_delivery": "Quantity Delivered 1",
            "second_qty_delivery": "Quantity Delivered 2",
            "third_qty_delivery": "Quantity Delivered 3",
            "first_qty_uom": "Quantity For First Delivery",
            "second_qty_uom": "Quantity For Second Delivery",
            "third_qty_uom": "Quantity For Third Delivery",
        }, inplace=True)
        print(1)
        by_request_items_df.rename(columns={
            "by_request_item__bos_code": "BOS Code",
            'by_request_item__category': "Category",
            "by_request_item__bos_material_description": "Material Description",
            "by_request_item__delivery_uom" : "Delivery UOM",
            "total_weekly_request": "Total Weekly Request",
            "first_delivery": "Delivery 1",
            "second_delivery": "Delivery 2",
            "third_delivery": "Delivery 3",
            "first_final_delivery": "Final Delivery 1",
            "second_final_delivery": "Final Delivery 2",
            "third_final_delivery": "Final Delivery 3",
            "first_qty_delivery": "Quantity Delivered 1",
            "second_qty_delivery": "Quantity Delivered 2",
            "third_qty_delivery": "Quantity Delivered 3",
            "first_qty_uom": "Quantity For First Delivery",
            "second_qty_byrequest_uom" : "Quantity For Second Delivery",
            "third_qty_byrequest_uom" : "Quantity For Third Delivery"
        }, inplace=True)

        print(2)
        delivery_selected = delivery_items_df[
            ["BOS Code", "Category","Delivery UOM","Material Description", "Quantity For First Delivery", "Quantity For Second Delivery","Quantity For Third Delivery",]
        ].copy()
        print(2.5)
        by_request_selected = by_request_items_df[
            ["BOS Code", "Category","Delivery UOM", "Material Description", "Quantity For First Delivery", "Quantity For Second Delivery","Quantity For Third Delivery"]
        ].copy()

        # add an identifier column to distinguish sections
        delivery_selected["Item Classification"] = "For Delivery Items"
        by_request_selected["Item Classification"] = "By Request"
        print(3)
        # create a boundary row

        boundary_row = pd.DataFrame({
            "BOS Code": ["BOS Code"],
            "Category": ["Category"],
            "Delivery UOM": ["Delivery UOM"],
            "Material Description": ["Material Description"],
            "Quantity For First Delivery": ["Quantity For First Delivery"],
            "Quantity For Second Delivery": ["Quantity For Second Delivery"],
            "Quantity For Third Delivery": ["Quantity For Third Delivery"],
            "Item Classification" : ["Item Classification"]

        })
        boundary_row_styled = boundary_row.style.applymap(lambda x: 'background-color: yellow')
        #combine DataFrames with the boundary
        final_df = pd.concat([delivery_selected, boundary_row, by_request_selected], ignore_index=True)

        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            if not delivery_items_df.empty:
                delivery_items_df.to_excel(writer, index=False, sheet_name="Delivery Items")
            if not by_request_items_df.empty:
                by_request_items_df.to_excel(writer, index=False, sheet_name="By Request Items")
            if not final_df.empty:
                final_df.to_excel(writer, index=False, sheet_name="Consolidated Reports")

        wb = openpyxl.load_workbook(file_path)
        sheet = wb["Consolidated Reports"]

        # style the boundary row
        for cell in sheet[delivery_selected.shape[0] + 2]:
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type = "solid")

        # save the workbook
        wb.save(file_path)

        logger.info(f"✅ Updated {file_name} with Delivery Items & By Request Items.")
        return file_name

    except Exception as e:
        logger.error(f"❌ Error updating sales report for {mrp.inventory_code}: {e}")


def fuzzy_match(query, choices, limit=3, threshold=70):
    matches = process.extract(query, choices, limit=limit, score_cutoff=threshold)
    return [(match, score) for match, score, _ in matches]


class UserDefinedVariablesListCreateView(ListCreateAPIView):
    queryset = UserDefinedVariables.objects.all()
    serializer_class = UserDefinedVariablesSerializer

class UserDefinedVariablesDetailView(RetrieveUpdateDestroyAPIView):
    queryset = UserDefinedVariables.objects.all()
    serializer_class = UserDefinedVariablesSerializer




class CustomTokenObtainPairView(TokenObtainPairView):

    def post(self, request, *args, **kwargs):
        username = request.data.get('username')
        password = request.data.get('password')


        user = User.objects.filter(username=username).first()
        if not user:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)


        try:
            employee = Employee.objects.get(user=user)
        except Employee.DoesNotExist:
            return Response({"error": "User not associated with a company"}, status=status.HTTP_400_BAD_REQUEST)

        #Check if the user is locked
        if employee.locked == 1:
            return Response({"error": "User is locked. Please contact administrator."},
                            status=status.HTTP_403_FORBIDDEN)

        authenticated_user = authenticate(username=username, password=password)
        if authenticated_user:

            employee.attempts = 0
            employee.save()

            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }, status=status.HTTP_200_OK)
        else:
            employee.attempts += 1
            employee.save()

            #lock the user after 3 failed attempts
            if employee.attempts >= 3:
                employee.locked = 1
                employee.save()
                return Response({"error": "User is locked due to too many failed login attempts."},
                                status=status.HTTP_403_FORBIDDEN)

            return Response({"error": "Invalid credentials"}, status=status.HTTP_401_UNAUTHORIZED)


class ModuleListView(ListAPIView):
    
    def get(self, request):
        modules = Modules.objects.all()
        serializer = ModuleSerializer(modules, many=True)
        return Response(serializer.data)


class EmployeeListView(ListCreateAPIView):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer

    def get(self, request, *args, **kwargs):

        draw = int(request.GET.get('draw', 1))
        page = int(request.GET.get('page', 1))
        length = int(request.GET.get('pageSize', 10))
        search_value = request.GET.get('search', '')
        order_column = request.GET.get('sortColumnIndex', "user__first_name")
        order_direction = request.GET.get('sortDirection', 'asc')
        offset = (page - 1) * length

        if order_direction == 'desc':
            order_column = f"-{order_column}"

        queryset = self.queryset
        if search_value:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_value) |
                Q(user__last_name__icontains=search_value) |
                Q(user__email__icontains=search_value) |
                Q(user__employee__id__icontains=search_value) |

                Q(user__employee__superior__user__first_name__icontains=search_value) |
                Q(user__employee__superior__user__last_name__icontains=search_value) |
                Q(user__email__icontains=search_value) |
                Q(user__last_login__icontains=search_value) |
                Q(user__employee__cellphone_number__icontains=search_value) |
                Q(user__is_active__icontains=search_value)
            ).distinct()

        for_pagination = Paginator(queryset, length)
        queryset = queryset.order_by(order_column)[offset:offset + (length)]
        serializer = self.serializer_class(queryset, many=True)
        response_data = {
            "draw": draw,
            "recordsTotal": self.queryset.count(),
            "recordsFiltered": 1,
            "page_count": for_pagination.count,
            "page_num_pages": for_pagination.num_pages,
            "data": serializer.data,
        }

        return Response(response_data, status=status.HTTP_200_OK)


    def create(self, request, *args, **kwargs):
        data = request.data

        try:
            generated_password = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
            hashed_password = make_password(generated_password)
            generated_username = f"{data.get('first_name')}{data.get('last_name')}"

            user_email = data.get("email")
            user_data = {
                "username": generated_username,
                "email": user_email,
                "first_name": data.get("first_name"),
                "last_name": data.get("last_name"),
                "password": hashed_password,
            }

            user, created = User.objects.get_or_create(username=user_data["username"], defaults=user_data)
            if not created:
                return Response({"error": "User with this email already exists."}, status=status.HTTP_400_BAD_REQUEST)

            employee = Employee.objects.create(
                user=user,
                department=Departments.objects.get(id=data["department"]),
                cellphone_number=data.get("mobile_number"),
                telephone_number=data.get("telephone_number"),
                superior=Employee.objects.filter(id=int(data.get("supervisor"))).first(),
                added_by=request.user,
            )

            if "role" in data:
                role = Roles.objects.get(id=int(data['role']))
                employee.role = role

            if "area" in data:
                areas = Area.objects.filter(location__in=[item['value'] for item in data['area']])
                employee.area.add(*areas)

            if "permissions" in data:
                permission_codenames = [actions for actions, has_permission in data["permissions"].items() if has_permission]

                permissions = ModulePermissions.objects.filter(codename__in=permission_codenames)
                employee.module_permissions.add(*permissions)


                def get_module_hierarchy(module):
                    hierarchy = []
                    current_module = module
                    while current_module:
                        hierarchy.append(current_module)
                        current_module = current_module.parent_module
                    return hierarchy

                modules_to_add = set()
                for permission in permissions:

                    module = permission.module
                    if module:
                        module_hierarchy = get_module_hierarchy(module)
                        modules_to_add.update(module_hierarchy)



                employee.modules.add(*modules_to_add)

            employee.save()

            # Email logic (optional - commented out)
            # send_mail(
            #     "New User",
            #     f"Hi. \nYour Password is {generated_password}",
            #     "your_email@example.com",
            #     [user_email],
            #     fail_silently=False,
            # )

            return_data = {
                "data": EmployeeSerializer(employee).data,
                "message": f"Default Password is {generated_password}. Please save it immediately as this window will close soon.",
            }
            return Response(return_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EmployeeDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
            serializer = UserDetailSerializer(user)
            return Response(serializer.data)

        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=404)



class CombinedModuleListView(APIView):
    def get(self, request):
        modules_with_no_submodules = Modules.objects.filter(submodules__isnull=True)
        submodules = Modules.objects.all()
        modules = list(modules_with_no_submodules) + list(submodules)
        serializer = ModuleSerializer(modules, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AreaListView(ListAPIView):

    def get(self, request):
        area = Area.objects.all()
        serializer = AreaSerializer(area, many=True)
        return Response(serializer.data)



class RoleListCreate(ListCreateAPIView):
    serializer_class = RoleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Roles.objects.all()

    def perform_create(self, serializer):
        role_instance = serializer.save()
        data = self.request.data
        if 'area' in data:
            areas = Area.objects.filter(location__in=[item['value'] for item in data['area']])
            role_instance.area.add(*areas)


        if "permissions" in data:
                permission_codenames = [actions for actions, has_permission in data["permissions"].items() if has_permission]
                permissions = ModulePermissions.objects.filter(codename__in=permission_codenames)
                role_instance.permissions.add(*permissions)

                def get_module_hierarchy(module):
                    hierarchy = []
                    current_module = module
                    while current_module:
                        hierarchy.append(current_module)
                        current_module = current_module.parent_module
                    return hierarchy

                modules_to_add = set()
                for permission in permissions:

                    module = permission.module
                    if module:
                        module_hierarchy = get_module_hierarchy(module)
                        modules_to_add.update(module_hierarchy)

                role_instance.modules.add(*modules_to_add)

        if 'copy_area' in data and data['copy_area']:
            for role_data in data['roles']:
                role = Roles.objects.get(id=role_data['id'])
                role.area.clear()
                role.area.add(*areas)

        if 'copy_permissions' in data and data['copy_permissions']:
            for role_data in data['roles']:
                role = Roles.objects.get(id=role_data['id'])
                role.permissions.clear()
                role.permissions.add(*permissions)

                modules_to_add_for_roles = set()
                for permission in permissions:
                    module = permission.module
                    if module:
                        module_hierarchy = get_module_hierarchy(module)
                        modules_to_add_for_roles.update(module_hierarchy)

                role.modules.clear()
                role.modules.add(*modules_to_add_for_roles)

        role_instance.save()



class EmployeeEditView(RetrieveUpdateAPIView):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer

    def update(self, request, *args, **kwargs):
        data = request.data

        try:
            employee = self.get_object()
            user = employee.user

            user.first_name = data.get("first_name", user.first_name)
            user.last_name = data.get("last_name", user.last_name)
            user.email = data.get("email", user.email)
            user.save()

            employee.department = Departments.objects.get(id=data.get("department", employee.department.id))

            employee.cellphone_number = data.get("mobile_number", employee.cellphone_number)

            employee.telephone_number = data.get("telephone_number", employee.telephone_number)

            employee.superior = Employee.objects.filter(id=data.get("supervisor", employee.superior.id)).first()


            # if "role" in data:
            #     role_id = data["role"]
            #     role = Roles.objects.get(id=role_id)
            #     employee.role = role
            #     print(1)

            if "area" in data:
                area_instances = Area.objects.filter(location__in=[area["value"] for area in data["area"]])
                employee.area.set(area_instances)


            if "permissions" in data:


                employee.module_permissions.clear()
                modules_to_add = set()
                for action, has_permission in data['permissions'].items():
                    if has_permission:
                        permission = ModulePermissions.objects.filter(codename=action).first()
                        if permission:
                            employee.module_permissions.add(permission)
                            def get_module_hierarchy(module):
                                hierarchy = []
                                while module:
                                    hierarchy.append(module)
                                    module = module.parent_module
                                return hierarchy
                            module_hierarchy = get_module_hierarchy(permission.module)
                            modules_to_add.update(module_hierarchy)

                employee.modules.set(modules_to_add)
            employee.save()
            return Response(EmployeeSerializer(employee).data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Failed to update employee. Details: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )


class RoleEditView(RetrieveUpdateAPIView):
    queryset = Roles.objects.all()
    serializer_class = RolesSerializerPlain

    def update(self, request, pk,*args, **kwargs):

        try:
            data = request.data
            role_instance = Roles.objects.get(id=pk)

            if "area" in data:
                role_instance.area.clear()
                areas = Area.objects.filter(location__in=[item['value'] for item in data['area']])
                role_instance.area.add(*areas)

            if "permissions" in data:
                role_instance.permissions.clear()

                modules_to_add = set()
                for action, has_permission in data['permissions'].items():

                    if has_permission:
                        permission = ModulePermissions.objects.filter(codename=action).first()
                        if permission:
                            role_instance.permissions.add(permission)

                            def get_module_hierarchy(module):
                                hierarchy = []
                                while module:
                                    hierarchy.append(module)
                                    module = module.parent_module

                                return hierarchy

                            module_hierarchy = get_module_hierarchy(permission.module)
                            modules_to_add.update(module_hierarchy)

                role_instance.modules.set(modules_to_add)

            role_instance.save()
            return Response(
                    RolesSerializerPlain(role_instance).data,
                    status=status.HTTP_200_OK
                )
        except Roles.DoesNotExist:
                return Response(
                    {"detail": "Role not found."},
                    status=status.HTTP_404_NOT_FOUND
                )


class CombinedDataView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        modules_with_no_submodules = Modules.objects.filter(submodules__isnull=True)
        module_serializer = ModulesSerializerParent(modules_with_no_submodules, many=True)

        areas = Area.objects.all().values()
        departments = Departments.objects.all()
        dept_serializer = DepartmentsSerializer(departments, many=True)
        roles = Roles.objects.all()
        role_serializer = RoleSerializer(roles, many=True)

        supervisor_role = roles.filter(role='Supervisors').first()
        if supervisor_role:
            employees_under_supervisor_role = Employee.objects.filter(role=supervisor_role)
            supervisor_data = employees_under_supervisor_role.values(
                'user__id', 'user__username', 'user__first_name', 'user__last_name'
            )
        else:
            supervisor_data = []
        response_data = {
            "modules": module_serializer.data,
            "areas": areas,
            "roles": role_serializer.data,
            "departments": dept_serializer.data,
            "supervisors": supervisor_data,
        }

        return Response(response_data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = RoleSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(user=request.user)  # Save the role with the current user
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data)

        if serializer.is_valid():
            user = request.user
            old_password = serializer.validated_data['old_password']
            new_password = serializer.validated_data['new_password']
            if not user.check_password(old_password):
                return Response({"new_password": ["You entered wrong password."]}, status=status.HTTP_400_BAD_REQUEST)

            user.set_password(new_password)
            user.save()
            return Response({"detail": "Password changed successfully."}, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class EmployeeFlatDetailView(APIView):

    def get(self, request, pk):
        try:
            employee = Employee.objects.get(pk=pk)
        except Employee.DoesNotExist:
            return Response({'detail': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        serializer = EmployeeSerializerPlain(employee)
        return Response(serializer.data)


class RoleFlatDetailView(APIView):

    def get(self, request, pk):
        try:
            role = Roles.objects.get(pk=pk)
        except Roles.DoesNotExist:
            return Response({'detail': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        serializer = RolesSerializerPlain(role)
        return Response(serializer.data)


class AccessKeyView(ListAPIView):

    def get(self, request):
        access_key = AccessKey.objects.all()
        serialized_key = AccessKeySerializer(access_key, many=True)

        return Response(serialized_key.data)



"""Start of Upload excel part."""



class PosItemsUploadView(APIView):
    def post(self, request):
        serializer = FileUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            try:
                df = pd.read_excel(file, engine='openpyxl')
                if 'POS CODE' not in df.columns or 'MENU - DESCRIPTION' not in df.columns:
                    return Response({"error": "Missing 'sku_code' or 'sku_description' columns."}, status=status.HTTP_400_BAD_REQUEST)

                distinct_skus = df[['POS CODE', 'MENU - DESCRIPTION']].drop_duplicates(subset='POS CODE')

                pos_items = [
                    PosItems(menu_description=row['MENU - DESCRIPTION'], pos_item=row['POS CODE'])
                    for _, row in distinct_skus.iterrows()
                ]

                with transaction.atomic():
                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_positems DISABLE KEYS;')

                    PosItems.objects.bulk_create(pos_items, batch_size=1000)

                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_positems ENABLE KEYS;')

                return Response({"message": "✅ POS Items imported successfully!"}, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UploadBOMMasterlist(APIView):
    def post(self, request):
        serializer = FileUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            try:
                df = pd.read_excel(file, engine='openpyxl')

                required_columns = [
                    'POS CODE', 'CATEGORY', 'PD ITEM DESCRIPTION', 'BOM', 'UOM',
                    'BOS CODE', 'BOS MATERIAL DESCRIPTION', 'BOS MATERIAL UOM'
                ]
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    return Response({"error": f"Missing columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

                df = df.where(pd.notna(df), None)

                pos_codes = df['POS CODE'].unique()
                bos_codes = df['BOS CODE'].unique()

                pos_items_map = {item.pos_item: item for item in PosItems.objects.filter(pos_item__in=pos_codes)}
                bos_items_map = {item.bos_code: item for item in BosItems.objects.filter(bos_code__in=bos_codes)}

                chunk_size = 1000
                num_chunks = int(np.ceil(len(df) / chunk_size))
                chunks = np.array_split(df, num_chunks)

                with transaction.atomic():
                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_bommasterlist DISABLE KEYS;')

                    for chunk in chunks:
                        bom_objects = []
                        for _, row in chunk.iterrows():
                            pos_item = pos_items_map.get(str(row['POS CODE']))
                            bos_item = bos_items_map.get(str(row['BOS CODE']))

                            if pos_item or bos_item:
                                bom_objects.append(
                                    BomMasterlist(
                                        pos_code=pos_item,
                                        bos_code=bos_item,  # Assign the instance, NOT a raw string
                                        category=row['CATEGORY'],
                                        item_description=row['PD ITEM DESCRIPTION'],
                                        bom=row['BOM'],
                                        uom=row['UOM'],
                                    )
                                )
                        BomMasterlist.objects.bulk_create(bom_objects, batch_size=1000)

                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_bommasterlist ENABLE KEYS;')

                return Response({"message": "✅ Data imported successfully!"}, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UploadByRequest(APIView):
    def post(self, request):
        serializer = FileUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            try:
                df = pd.read_excel(file, engine='openpyxl', sheet_name='BY REQUEST', header=6)

                required_columns = ['BOS MATCODE', 'BOS MATERIAL DESCRIPTION', 'BOS UOM', 'COLD/DRY/FOR PR', 'DELIVERY UOM']
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    return Response({"error": f"Missing columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)


                df = df.where(pd.notna(df), None)
                df['BOS MATCODE'] = df['BOS MATCODE'].astype(str)

                existing_bos_codes = set(ByRequestItems.objects.filter(bos_code__in=df['BOS MATCODE']).values_list('bos_code', flat=True))


                new_entries = []
                for _, row in df.iterrows():
                    if row['BOS MATCODE'] not in existing_bos_codes:  # Avoid inserting duplicates
                        new_entries.append(
                            ByRequestItems(
                                bos_code=row['BOS MATCODE'],
                                bos_material_description=row['BOS MATERIAL DESCRIPTION'],
                                bos_uom=row['BOS UOM'],
                                category=row['COLD/DRY/FOR PR'],
                                delivery_uom=row['DELIVERY UOM'],
                            )
                        )


                chunk_size = 1000
                with transaction.atomic():
                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_byrequestitems DISABLE KEYS;')

                    for i in range(0, len(new_entries), chunk_size):
                        ByRequestItems.objects.bulk_create(new_entries[i:i + chunk_size], batch_size=chunk_size)

                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_byrequestitems ENABLE KEYS;')

                return Response({"message": f"✅ Imported {len(new_entries)} new records!"}, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
# orig 294933
# Last Row 386070
# 335874

#1028209 109402
#1214673 88019

class SalesUploadView(APIView):
    def get_file_hash(self, file):
        """Generates a SHA-256 hash for the uploaded file."""
        hasher = hashlib.sha256()
        for chunk in file.chunks():
            hasher.update(chunk)
        return hasher.hexdigest()

    def post(self, request):
        serializer = FileUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']

            file_hash = self.get_file_hash(file)
            if UploadedFile.objects.filter(file_hash=file_hash).exists():
                return Response({"error": "🚫 This file has already been uploaded."}, status=status.HTTP_400_BAD_REQUEST)

            try:
                df = pd.read_excel(file, engine='openpyxl')
                df.columns = df.columns.str.strip().str.upper()  # Normalize column names

                required_columns = {
                    'IFS CODE', 'NAME OF OUTLET', 'OR NO.', 'CUSTOMER NAME', 'SKU CODE',
                    'QTY', 'UNIT PRICE', 'GROSS SALES', 'TYPE OF DISCOUNT', 'DISC AMOUNT',
                    'VAT DEDUCT', 'NET SALES', 'MODE OF PAYMENT', 'TRANSACTION TYPE',
                    'NOTE', 'REMARKS', 'SALES DATE', 'TIME'
                }


                missing_columns = required_columns - set(df.columns)
                if missing_columns:
                    return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"},
                                    status=status.HTTP_400_BAD_REQUEST)


                area_dict = {a.location: a for a in Area.objects.all()}
                pos_items_dict = {p.pos_item: p for p in PosItems.objects.all()}

                sales_objects = []
                for _, row in df.iterrows():
                    try:

                        outlet = area_dict.get(row['NAME OF OUTLET'])
                        if not outlet:
                            outlet, _ = Area.objects.get_or_create(location=row['NAME OF OUTLET'])
                            area_dict[row['NAME OF OUTLET']] = outlet


                        pos_item = pos_items_dict.get(row['SKU CODE'])
                        if not pos_item:
                            pos_item, _ = PosItems.objects.get_or_create(pos_item=row['SKU CODE'])
                            pos_items_dict[row['SKU CODE']] = pos_item


                        sales_objects.append(
                            Sales(
                                ifs_code=row['IFS CODE'],
                                outlet=outlet,
                                or_number=row['OR NO.'],
                                customer_name=row.get('CUSTOMER NAME'),
                                sku_code=pos_item,
                                quantity=row['QTY'],
                                unit_price=row['UNIT PRICE'],
                                gross_sales=row['GROSS SALES'],
                                type_of_discount=row.get('TYPE OF DISCOUNT'),
                                discount_amount=row['DISC AMOUNT'],
                                vat_deduct=row['VAT DEDUCT'],
                                net_sales=row['NET SALES'],
                                mode_of_payment=row.get('MODE OF PAYMENT'),
                                transaction_type=row['TRANSACTION TYPE'],
                                note=row.get('NOTE'),
                                remarks=row.get('REMARKS'),
                                sales_date=row['SALES DATE'],
                                time=row['TIME']
                            )
                        )

                    except Exception as e:
                        print(f"⚠️ Skipping row due to error: {e}")


                with transaction.atomic():
                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_sales DISABLE KEYS;')

                    Sales.objects.bulk_create(sales_objects, batch_size=1000)

                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_sales ENABLE KEYS;')


                    UploadedFile.objects.create(file_hash=file_hash)

                return Response({"message": "✅ Sales data imported successfully!"}, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class EndingInventoryUploadView(APIView):
    def post(self, request):

        serializer = FileUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            area_id = request.data['area_id']

            if not area_id:
                return Response({"error": "Missing 'area_id' in request."}, status=status.HTTP_400_BAD_REQUEST)

            area = Area.objects.filter(id=area_id).first()
            if not area:
                return Response({"error": f"Area ID {area_id} not found."}, status=status.HTTP_400_BAD_REQUEST)
            if InventoryCode.objects.filter(area=area, status__lt=5).exists():
                return Response(
                    {"error": "Cannot create inventory code. Some inventory codes in this area have a pending status."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            inventory_code = InventoryCode.objects.create(area=area, uploaded_by=request.user)

            try:
                df = pd.read_excel(file, engine='openpyxl')
                required_columns = ["BOS MATCODE", "BOS MATERIAL DESCRIPTION", "QTY", "COLD/DRY/FOR PR"]
                if not all(col in df.columns for col in required_columns):
                    return Response({"error": "Missing required columns."}, status=status.HTTP_400_BAD_REQUEST)

                distinct_entries = df[["BOS MATCODE", "BOS MATERIAL DESCRIPTION", "QTY", "COLD/DRY/FOR PR"]].drop_duplicates(subset="BOS MATCODE")

                inventory_items = []

                for _, row in distinct_entries.iterrows():
                    try:
                        bom_entry = BosItems.objects.filter(bos_code=row["BOS MATCODE"]).first()

                        if not bom_entry:
                            bom_entry = BosItems.objects.create(
                                bos_code=row["BOS MATCODE"],
                                bos_material_description=row["BOS MATERIAL DESCRIPTION"]
                            )

                        inventory_items.append(EndingInventory(
                            inventory_code=inventory_code,
                            bom_entry=bom_entry,

                            actual_ending=row["QTY"] if pd.notna(row["QTY"]) else 0
                        ))

                    except Exception as e:
                        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

                with transaction.atomic():
                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_endinginventory DISABLE KEYS;')

                    EndingInventory.objects.bulk_create(inventory_items, batch_size=1000)

                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_endinginventory ENABLE KEYS;')

                return Response({"message": "✅ Ending Inventory imported successfully!"}, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class BosItemsUploadView(APIView):
    def post(self, request):
        serializer = FileUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            try:
                df = pd.read_excel(file, engine='openpyxl')
                required_columns = {
                    "BOS MATCODE": "bos_code",
                    "BOS MATERIAL DESCRIPTION": "bos_material_description",
                    "BOS UOM": "bos_uom",
                    "DELIVERY UOM": "delivery_uom",
                    "BUNDLING SIZE": "bundling_size",
                    "CONVERSION DELIVERY UOM": "conversion_delivery_uom",
                    "COLD/DRY/FOR PR": "category",
                }

                missing_columns = [col for col in required_columns.keys() if col not in df.columns]
                if missing_columns:
                    return Response({"error": f"Missing columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)


                df.rename(columns=required_columns, inplace=True)
                df.drop_duplicates(subset=['bos_code'], inplace=True)
                df = df.where(pd.notna(df), None)

                #prepare for bulk insert data
                bos_items = [
                    BosItems(
                        bos_code=row['bos_code'],
                        bos_material_description=row['bos_material_description'],
                        bos_uom=row['bos_uom'],
                        delivery_uom=row['delivery_uom'],
                        bundling_size=row['bundling_size'],
                        conversion_delivery_uom=row['conversion_delivery_uom'],
                        category=row['category'],
                    ) for _, row in df.iterrows()
                ]

                #use bulk_create for efficient inserts
                with transaction.atomic():
                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_bositems DISABLE KEYS;')

                    BosItems.objects.bulk_create(bos_items, batch_size=1000)

                    with connection.cursor() as cursor:
                        cursor.execute('ALTER TABLE mrp_api_bositems ENABLE KEYS;')

                return Response({"message": "✅ BOS Items imported successfully!"}, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

"""End of upload part"""

class InventoryCodeByAreaView(ListAPIView):
    serializer_class = InventoryCodeSerializer

    def get_queryset(self):
        area_id = self.kwargs.get("area_id")
        return InventoryCode.objects.filter(area_id=area_id)

# class InventoryCodeByAreaView(APIView):
#     def get(self, request, pk):
#         try:
#             area = Area.objects.get(id=pk)
#             inventory_codes = InventoryCode.objects.filter(area=area)
#             serializer = InventoryCodeSerializer(inventory_codes, many=True)
#             return Response(serializer.data)
#         except Area.DoesNotExist:
#             return Response({"detail": "Area not found"}, status=status.HTTP_404_NOT_FOUND)


class ForecastByInventoryCodeView(APIView):
    def get(self, request, pk):
        try:
            inventory_code = InventoryCode.objects.get(id=pk)
            forecasts = Forecast.objects.filter(inventory_code=inventory_code)
            serializer = ForecastSerializer(forecasts, many=True)
            return Response(serializer.data)
        except InventoryCode.DoesNotExist:
            return Response({"detail": "Inventory code not found"}, status=status.HTTP_404_NOT_FOUND)



class InsertDeliveryItemsView(APIView):

    def post(self, request, pk):
        data = request.data
        for x in data['by_request_items']:
            print(x)
            print()


        try:
            inventory_code = InventoryCode.objects.get(id=pk)
        except InventoryCode.DoesNotExist:
            return Response({"detail": f"InventoryCode with ID {pk} not found."}, status=status.HTTP_404_NOT_FOUND)

        delivery_code = DeliveryCode.objects.create(
            inventory_code=inventory_code,
            requested_by=request.user
        )

        bom_entry_ids = {item.get('bom_entry__id') for item in data.get('adjustment', [])}
        bom_entries = BosItems.objects.in_bulk(bom_entry_ids)  # Fetch all at once
        missing_bom_entries = bom_entry_ids - set(bom_entries.keys())

        if missing_bom_entries:
            return Response(
                {"detail": f"BosItems with IDs {list(missing_bom_entries)} not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        delivery_items_objects = [
            DeliveryItems(
                delivery_code=delivery_code,
                bom_entry=bom_entries[item['bom_entry__id']],  # Use pre-fetched object
                first_adjustment = item['first_adjustment'],
                second_adjustment = item['second_adjustment'],
                third_adjustment = item['third_adjustment'],
                first_final_delivery = item['first_final_delivery'],
                second_final_delivery =  item['second_final_delivery'],
                third_final_delivery = item['third_final_delivery'],
                first_qty_delivery =  item['first_qty_delivery'],
                second_qty_delivery = item['second_qty_delivery'],
                third_qty_delivery =  item['third_qty_delivery'],
                first_qty_uom =  f"{round(item['first_qty_delivery'],2)} {bom_entries[item['bom_entry__id']].bos_uom}" if item['first_qty_delivery'] else "-",
                second_qty_uom =  f"{round(item['second_qty_delivery'],2)} {bom_entries[item['bom_entry__id']].bos_uom}" if item['second_qty_delivery'] else "-",
                third_qty_uom =  f"{round(item['third_qty_delivery'],2)} {bom_entries[item['bom_entry__id']].bos_uom}" if item['third_qty_delivery'] else "-",



            )
            for item in data.get('adjustment', [])
        ]

        if delivery_items_objects:
            DeliveryItems.objects.bulk_create(delivery_items_objects)

        if data['by_request_items']:

            by_request_objects = []
            by_request_ids = {by_request['by_request_item'] for by_request in data.get('by_request_items', [])}
            by_request_items = ByRequestItems.objects.in_bulk(by_request_ids)

            missing_by_requests = by_request_ids - set(by_request_items.keys())
            if missing_by_requests:
                return Response(
                    {"detail": f"ByRequestItems with IDs {list(missing_by_requests)} not found."},
                    status=status.HTTP_404_NOT_FOUND
                )

            for by_request in data.get('by_request_items', []):
                by_request_item = by_request_items[by_request['by_request_item']]
                by_request_objects.append(
                    ByRequest(
                        delivery_code=delivery_code,
                        by_request_item=by_request_item,
                        total_weekly_request=by_request['first_delivery'] + by_request['second_delivery'] + by_request['third_delivery'],
                        first_delivery=by_request['first_delivery'],
                        second_delivery=by_request['second_delivery'],
                        third_delivery=by_request['third_delivery'],
                        first_qty_delivery=by_request['first_delivery'] * by_request_item.conversion,
                        second_qty_delivery=by_request['second_delivery'] * by_request_item.conversion,
                        third_qty_delivery=by_request['third_delivery'] * by_request_item.conversion,
                        first_qty_uom=f'{round(by_request['first_delivery'] * by_request_item.conversion,2)} {by_request_item.bos_uom}' if by_request['first_delivery'] else "-",
                        second_qty_byrequest_uom=f'{round(by_request['first_delivery'] * by_request_item.conversion,2)} {by_request_item.bos_uom}' if by_request['first_delivery'] else "-",
                        third_qty_byrequest_uom=f'{round(by_request['first_delivery'] * by_request_item.conversion,2)} {by_request_item.bos_uom}' if by_request['first_delivery'] else "-",

                    )
                )

            if by_request_objects:
                ByRequest.objects.bulk_create(by_request_objects)
        draft_status = Status.objects.get(id=3)
        inventory_code.status = draft_status
        inventory_code.number_of_request = data['number_of_request']
        inventory_code.number_of_items = data['number_of_items']
        inventory_code.save()
        return Response(
            {"inserted_delivery_items": DeliveryItemsSerializer(delivery_items_objects, many=True).data},
            status=status.HTTP_201_CREATED
        )

class UpdateDeliveryItemsView(APIView):

    def post(self, request, pk):
        data = request.data
        edit_message = data.get('edit_message')


        try:
            inventory_code = InventoryCode.objects.get(id=pk)
        except InventoryCode.DoesNotExist:
            return Response({"detail": f"InventoryCode with ID {pk} not found."}, status=status.HTTP_404_NOT_FOUND)


        delivery_code = DeliveryCode.objects.get(inventory_code=inventory_code)
        delivery_code.updated_by = request.user
        delivery_code.updated_at = timezone.now()
        delivery_code.save()

        bom_entry_ids = {item.get('bom_entry__id') for item in data.get('adjustment', [])}
        bom_entries = BosItems.objects.in_bulk(bom_entry_ids)

        missing_bom_entries = bom_entry_ids - set(bom_entries.keys())
        if missing_bom_entries:
            return Response(
                {"detail": f"BosItems with IDs {list(missing_bom_entries)} not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        delivery_items_objects = []
        for item in data.get('adjustment', []):
            delivery_item, created = DeliveryItems.objects.update_or_create(
                delivery_code=delivery_code,

                bom_entry=bom_entries[item['bom_entry__id']],  # Use pre-fetched object
                defaults={
                    'first_adjustment' : item['first_adjustment'],
                    'second_adjustment' : item['second_adjustment'],
                    'third_adjustment' : item['third_adjustment'],
                    'first_final_delivery' : item['first_final_delivery'],
                    'second_final_delivery' :  item['second_final_delivery'],
                    'third_final_delivery' : item['third_final_delivery'],
                    'first_qty_delivery' :  item['first_qty_delivery'],
                    'second_qty_delivery' : item['second_qty_delivery'],
                    'third_qty_delivery' :  item['third_qty_delivery'],
                    'first_qty_uom' :  f"{round(item['first_qty_delivery'],2)} {bom_entries[item['bom_entry__id']].bos_uom}" if item['first_qty_delivery'] else "-",
                    'second_qty_uom' :  f"{round(item['second_qty_delivery'],2)} {bom_entries[item['bom_entry__id']].bos_uom}" if item['second_qty_delivery'] else "-",
                    'third_qty_uom' :  f"{round(item['third_qty_delivery'],2)} {bom_entries[item['bom_entry__id']].bos_uom}" if item['third_qty_delivery'] else "-",
                },

            )
            delivery_items_objects.append(delivery_item)


        if data['by_request_items']:
            by_request_ids = {by_request['by_request_item'] for by_request in data.get('by_request_items', [])}
            by_request_items = ByRequestItems.objects.in_bulk(by_request_ids)

            missing_by_requests = by_request_ids - set(by_request_items.keys())
            if missing_by_requests:
                return Response(
                    {"detail": f"ByRequestItems with IDs {list(missing_by_requests)} not found."},
                    status=status.HTTP_404_NOT_FOUND
                )

            for by_request in data.get('by_request_items', []):
                by_request_item = by_request_items[by_request['by_request_item']]
                ByRequest.objects.update_or_create(
                    delivery_code=delivery_code,

                    by_request_item=by_request_item,
                    defaults={
                        'total_weekly_request': by_request['first_delivery'] + by_request['second_delivery'] + by_request['third_delivery'],
                        'first_delivery': by_request['first_delivery'],
                        'second_delivery': by_request['second_delivery'],
                        'third_delivery': by_request['third_delivery'],
                        'first_qty_delivery': by_request['first_delivery'] * by_request_item.conversion,
                        'second_qty_delivery': by_request['second_delivery'] * by_request_item.conversion,
                        'third_qty_delivery': by_request['third_delivery'] * by_request_item.conversion,
                        'first_qty_uom': f'{round(by_request['first_delivery'] * by_request_item.conversion,2)} {by_request_item.bos_uom}' if by_request['first_delivery'] else "-",

                        'second_qty_byrequest_uom': f'{round(by_request['second_delivery'] * by_request_item.conversion,2)} {by_request_item.bos_uom}' if by_request['second_delivery'] else "-",

                        'third_qty_byrequest_uom': f'{round(by_request['third_delivery'] * by_request_item.conversion,2)} {by_request_item.bos_uom}' if by_request['third_delivery'] else "-",
                    }
                )
        inventory_code.number_of_request = data['number_of_request']
        inventory_code.number_of_items = data['number_of_items']
        inventory_code.save()

        if edit_message:
            subject = f"MRP Emergency Adjustment Update - Inventory ID {pk}"
            message = f"""
            An emergency adjustment has been made.

            Details:
            {edit_message}

            Updated by: {request.user.username}
            Timestamp: {timezone.now()}
            """

            recipient_list = ["faustnizzane@gmail.com"]  # Replace with actual recipient(s)

            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,  # Ensure this is set in settings.py
                recipient_list,
                fail_silently=False,
            )

        return Response(
            {"updated_delivery_items": DeliveryItemsSerializer(delivery_items_objects, many=True).data},
            status=status.HTTP_200_OK
        )


class DeleteDeliveryItemsView(APIView):
    def delete(self, request, pk):

        try:
            delivery_code = DeliveryCode.objects.get(id=pk)
        except DeliveryCode.DoesNotExist:
            return Response({"detail": f"DeliveryCode with ID {pk} not found."}, status=status.HTTP_404_NOT_FOUND)
        delivery_code.status_id = 4
        delivery_code.save()

        return Response({"message": f"DeliveryCode ID {pk} status changed to 4 (deleted)."}, status=status.HTTP_200_OK)



class UserAreasView(APIView):
    permission_classes = [IsAuthenticated]  # Require authentication

    def get(self, request,pk):
        if not pk:
            return Response({"error": "User ID is required"}, status=400)

        try:
            employee = Employee.objects.get(user_id=pk)  # Get Employee linked to user
        except Employee.DoesNotExist:
            return Response({"error": "Employee profile not found"}, status=404)

        user_areas = employee.area.all().values("id", "location")  # Get assigned areas
        return Response({"areas": list(user_areas)})



class InventoryCodeListView(ListAPIView):
    serializer_class = InventoryCodeSerializer

    def get_queryset(self):
        area_id = self.kwargs.get("area_id")
        return InventoryCode.objects.filter(area_id=area_id)


class InventoryCodeDetailView(RetrieveAPIView):
    serializer_class = InventoryCodeSerializer

    def get(self, request, *args, **kwargs):
        inventory_id = self.kwargs.get("pk")
        inventory_code = get_object_or_404(InventoryCode, id=inventory_id)
        serializer = self.get_serializer(inventory_code)


        user_def_vars = UserDefinedVariables.objects.filter(area=inventory_code.area).first()
        if inventory_code.status.status < 5:
            number_of_request = user_def_vars.number_of_request if user_def_vars else 1
            number_of_items = user_def_vars.number_of_items if user_def_vars else 1
        else:
            number_of_request = inventory_code.number_of_request if inventory_code else 1
            number_of_items = inventory_code.number_of_items if inventory_code else 1


        first_delivery_multiplier = user_def_vars.first_delivery_multiplier if user_def_vars else 1
        second_delivery_multiplier = user_def_vars.second_delivery_multiplier if user_def_vars else 0
        third_delivery_multiplier = user_def_vars.third_delivery_multiplier if user_def_vars else 0


        data = serializer.data
        data["number_of_request"] = number_of_request
        data["number_of_items"] = number_of_items
        data["delivery_multiplier"] = [first_delivery_multiplier, second_delivery_multiplier, third_delivery_multiplier]



        return Response(data)


class EndingInventoryListView(ListAPIView):
    serializer_class = EndingInventorySerializer


    def get_queryset(self):
        inventory_id = self.kwargs.get("pk")  # Get inventory_code ID from URL
        return EndingInventory.objects.filter(inventory_code__id=inventory_id)  # Filter records


class SalesReportListView(ListAPIView):
    serializer_class = SalesReportSerializer

    def get_queryset(self):
        inventory_id = self.kwargs['inventory_id']

        latest_sales = SalesReport.objects.filter(inventory_code_id=inventory_id) \
            .annotate(row_num=Window(
                expression=RowNumber(),
                partition_by=[F('pos_item_id')],
                order_by=F('created_at').desc()
            )) \
            .filter(row_num=1) \
            .select_related('pos_item', 'area', 'inventory_code')

        return latest_sales



class InitialReplenishmentListView(ListAPIView):
    def list(self, request, *args, **kwargs):
        inventory_id = self.kwargs['inventory_id']

        area_id = InventoryCode.objects.get(id=inventory_id).area_id
        start_date = request.GET.get('start_date', '2025-02-26 00:00:00')
        end_date = request.GET.get('end_date', '2025-02-27 00:00:00')

        query = """
            SELECT sr.sales_report_name,sr.created_at,bom.category, bom.bos_code_id,bom.item_description, pos.menu_description as pos_menu_description, pos.pos_item,  ini.*
            FROM mrp_api_salesreport as sr
            join mrp_api_initialreplenishment as ini on sr.id = ini.sales_report_id
            left join mrp_api_bommasterlist as bom on ini.bom_entry_id = bom.id
            left join mrp_api_positems as pos on sr.pos_item_id = pos.id
            left join mrp_api_inventorycode ic on sr.area_id = ic.inventory_code
            where ini.inventory_code_id = %s
            ORDER BY pos.pos_item,daily_sales;

        """

        params = [inventory_id]

        # Execute the query with parameters
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            result = cursor.fetchall()

        return JsonResponse(list(result), safe=False)


class ForecastListView(ListAPIView):
    def list(self, request, *args, **kwargs):
        inventory_code_id = self.kwargs.get('inventory_code_id')

        if inventory_code_id:
            forecast_data = list(Forecast.objects.select_related('bom_entry')
                .filter(inventory_code_id=inventory_code_id)
                .values(
                    'bom_entry__id',
                    'bom_entry__bos_code',
                    'bom_entry__bos_uom',
                    'bom_entry__bos_material_description',
                    'bom_entry__bundling_size',
                    'bom_entry__conversion_delivery_uom',
                    'id',
                    'inventory_code_id',
                    'average_daily_usage',
                    'days_to_last',
                    'forecast_weekly_consumption',
                    'forecasted_ending_inventory',
                    'converted_ending_inventory',
                    'forecast',


                )
            )

            delivery_adjustments = {
                item['bom_entry_id']: item
                for item in DeliveryItems.objects.filter(
                    delivery_code__inventory_code_id=inventory_code_id
                ).values(
                    'bom_entry_id',
                    'first_adjustment',
                    'first_final_delivery',
                    'first_qty_uom',
                    'second_adjustment',
                    'second_final_delivery',
                    'second_qty_uom',
                    'third_adjustment',
                    'third_final_delivery',
                    'third_qty_uom',

                )
            }

            for item in forecast_data:
                bom_entry_id = item['bom_entry__id']
                delivery_data = delivery_adjustments.get(bom_entry_id, {})
                item['first_adjustment'] = delivery_data.get('first_adjustment', 0)
                item['first_final_delivery'] = delivery_data.get('first_final_delivery', 0)
                item['first_qty_uom'] = delivery_data.get('first_qty_uom', 0)
                item['second_adjustment'] = delivery_data.get('second_adjustment', 0)
                item['second_final_delivery'] = delivery_data.get('second_final_delivery', 0)
                item['second_qty_uom'] = delivery_data.get('second_qty_uom', 0)
                item['third_adjustment'] = delivery_data.get('third_adjustment', 0)
                item['third_final_delivery'] = delivery_data.get('third_final_delivery', 0)
                item['third_qty_uom'] = delivery_data.get('third_qty_uom', 0)


            return JsonResponse(forecast_data, safe=False)

        return JsonResponse({"error": "Missing inventory_code_id"}, status=400)






class ByRequestItemsListView(ListAPIView):
    def get(self, request, *args, **kwargs):
        inventory_id = self.kwargs.get("inventory_id")
        try:
            inventory_code = InventoryCode.objects.get(id=inventory_id)
            delivery_code = DeliveryCode.objects.filter(inventory_code=inventory_code).first()

            if delivery_code:
                by_request_items = ByRequest.objects.filter(delivery_code=delivery_code)
                if by_request_items.exists():
                    serializer = ByRequestSerializerC(by_request_items, many=True)
                else:
                    by_request_items = ByRequestItems.objects.all()
                    serializer = ByRequestItemsSerializer(by_request_items, many=True)

            else:
                by_request_items = ByRequestItems.objects.all()
                serializer = ByRequestItemsSerializer(by_request_items, many=True)

            return Response(serializer.data)

        except InventoryCode.DoesNotExist:
            return Response([])  # Return empty list if inventory_id is invalid


class SubmitInventoryView(APIView):
    def post(self, request, idofinventory):
        requested_status = Status.objects.get(id=4)

        try:
            inventory = InventoryCode.objects.get(id=idofinventory)
            employees = Employee.objects.filter(
                area=inventory.area,
                module_permissions__codename="approve_inventory"
            ).distinct()

            # Get corresponding users
            users = User.objects.filter(employee__in=employees)
            emails = employees.values_list("user__email", flat=True)

            # Send an email
            subject = "Inventory Approval Required"
            message = f"Dear Approver,\n\nAn inventory in {inventory.area.location} requires your approval.\n\nPlease review it at your earliest convenience.\nhttp://localhost:5173/inventory/\n\nBest,\nInventory Team"

            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,  # Make sure this is configured in settings.py
                list(emails),  # Convert QuerySet to a list
                fail_silently=False,
            )


            inventory.status = requested_status
            inventory.save()

            return Response(
                {"message": "Inventory submitted successfully!"},
                status=status.HTTP_200_OK,
            )

        except InventoryCode.DoesNotExist:
            return Response(
                {"error": "Inventory not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class SalesReportListViewDL(ListAPIView):
    serializer_class = SalesReportSerializer

    def get_queryset(self):
        inventory_id = self.kwargs.get('inventory_id')
        return SalesReport.objects.filter(inventory_code_id=inventory_id)



class ConsolidatedItemsView(APIView):
    def get(self, request, inventory_id):
        # Query for Delivery Items
        delivery_query = (
            DeliveryItems.objects
            .select_related("delivery_code", "bom_entry")
            .filter(delivery_code__inventory_code_id=148)
            .annotate(
                bos_code=F("bom_entry__bos_code"),
                delivery_uom=F("bom_entry__delivery_uom"),
                bos_material_description=F("bom_entry__bos_material_description"),
                bos_uom=F("bom_entry__bos_uom"),
                source=Value("Delivery", output_field=CharField())  # Distinguish source
            )
            .values(
                "delivery_code_id",
                "bos_code",
                "bos_material_description",
                "delivery_uom",
                "bos_uom",
                "first_qty_uom",
                "second_qty_uom",
                "third_qty_uom",
                "source"
            )
        )

        # Query for ByRequest Items
        by_request_query = (
            ByRequest.objects
            .select_related("delivery_code", "by_request_item")
            .filter(delivery_code__inventory_code_id=148)
            .annotate(
                bos_code=F("by_request_item__bos_code"),
                delivery_uom=F("by_request_item__delivery_uom"),
                bos_material_description=F("by_request_item__bos_material_description"),
                bos_uom=F("by_request_item__bos_uom"),
                source=Value("ByRequest", output_field=CharField())  # Distinguish source
            )
            .values(
                "delivery_code_id",
                "bos_code",
                "bos_material_description",
                "delivery_uom",
                "bos_uom",
                "first_qty_uom",
                "second_qty_byrequest_uom",  # Adjusted for ByRequest
                "third_qty_byrequest_uom",
                "source"
            )
        )

        # Combine both queries using UNION
        combined_query = delivery_query.union(by_request_query)

        return Response(combined_query)
